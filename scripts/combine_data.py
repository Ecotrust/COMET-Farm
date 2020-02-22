# import system modules
import os, sys, csv
import ipdb

# import functions for extracting data from excel
from openpyxl import load_workbook

# check if argument for workbook has been given
if len(sys.argv) < 2:
    print("\npython3 script combine_data.py <GIS data location> <spreadsheet location>\n")
    print("Command-line arguments are as follows:\n")
    print("  <GIS data location> system location of comma separated data from GIS\n")
    print("    eg /usr/local/name/comet/data.csv\n")
    print("  <spreadsheet locatiion> system location of spreadsheet to add GIS data\n")
    print("    eg /usr/local/name/comet/data.xlsx\n\n")
    exit()

gis_dir =  sys.argv[1]
wb_dir = sys.argv[2]

wb = load_workbook(filename = wb_dir)
scenario_sheet = wb['scenario']

gis_values = {}

# create sheet for list of processed fields
processed_sheet = wb.create_sheet('processed')

# open GIS data file and save it as dict
with open(gis_dir) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    csv_dict = csv.DictReader(csv_file)
    gis_values = csv_dict

    # for row in gis_values:
        # ipdb.set_trace()
        # print(dict(row))
        # param  = scenario_sheet['A' + str(row)].value
        # param_val  = scenario_sheet['B' + str(row)].value
        # scenario_values.setdefault(param, param_val)

    # add gis data to spreadsheet
    for row in gis_values:
        field_sheet = wb.copy_worksheet(scenario_sheet)
        field_sheet.title = 'ready_' + row['field_ID']
        processed_sheet.append(["name", field_sheet.title])
        # for rowNum in range(1, field_sheet.max_row):

        field_sheet.cell(row=7, column=2).value = field_sheet.cell(row=7, column=2).value + row['CcopName'] + '_' + '_' + row['field_ID']
        field_sheet.cell(row=8, column=2).value = row['CRP_NUM'] #dc added 1/16/20
        field_sheet.cell(row=9, column=2).value = row['field_ID']
        field_sheet.cell(row=10, column=2).value = row['GEOM']
        field_sheet.cell(row=11, column=2).value = row['AREA']
        field_sheet.cell(row=12, column=2).value = '4326'

                # rowName = rowName.lower()
                # if rowName == 'id':
                    # field_sheet.cell(row=rowNum, column=2).value = row['field_ID']
                # if rowName == 'geom':
                #     field_sheet.cell(row=rowNum, column=2).value = 'Polygon' + row['GEOM'] #Dropped parens to eliminate double parens
                # if rowName == 'area':
                #     field_sheet.cell(row=rowNum, column=2).value = row['AREA'] #changed 'acres' to 'AREA'
                # if rowName == 'srid':
                #     field_sheet.cell(row=rowNum, column=2).value = '4326'
                # if rowName == 'crop_scenario_name':
                    # field_sheet.cell(row=rowNum, column=2).value = row['CcopName'] + '_' + field_sheet.cell(row=rowNum, column=2).value + '_' + row['field_ID']
                    #Original: field_sheet.cell(row=rowNum, column=2).value = 'Todo_polygon_id' + row['field_ID'] + 'Todo_scenario_id'  Old Version
                # if rowName == 'ccop_name':
                #     field_sheet.cell(row=rowNum, column=2).value = row['CcopName'] #CcopName
                # if rowName == 'crop_number':
                #     field_sheet.cell(row=rowNum, column=2).value = row['CRP_NUM'] #dc added 1/16/20
                # if rowName == 'planting_date':
                #     field_sheet.cell(row=rowNum, column=2).value = row['planting_date']  #dc added 1/16/20
                # if rowName == 'harvest_date':
                #     field_sheet.cell(row=rowNum, column=2).value = row['harvest_date'] #dc added 1/16/20
                # if rowName == 'tillage_date':
                #     field_sheet.cell(row=rowNum, column=2).value = row['till_date']    #dc added 1/16/20
                # if rowName == 'n_application_date':
                #     field_sheet.cell(row=rowNum, column=2).value = row['n_app_date'] #dc added 1/16/20
                # if rowName == 'grain':
                #     field_sheet.cell(row=rowNum, column=2).value = row['grain'] #dc added 1/16/20
                # if rowName == 'pre_80':
                #     pre_1980_sheet = wb['pre1980']
                #     pre_80_gis_val = row['pre_80'] #changed 'pre-80' to 'pre_80'
                #     pre_1980_val = pre_1980_sheet.cell(row=int(pre_80_gis_val), column=1).value
                #     field_sheet.cell(row=rowNum, column=2).value = pre_1980_val
                # if rowName == 'yr80_2000':
                #     year80_sheet = wb['yr80']
                #     year80_sheet_val = row['yr80_2000'] #changed 'Year1980-2000' to 'yr80_2000'
                #     yr_80_val = year80_sheet.cell(row=int(year80_sheet_val), column=1).value
                #     field_sheet.cell(row=rowNum, column=2).value = yr_80_val
                #if rowName == 'till80_200':
                    #till_80_sheet = wb['tillage']
                    #till_80_sheet_val = row['Year1980-2000_Tillage']
                    #till_80_val = till_80_sheet.cell(row=int(till_80_sheet_val), column=1).value
                    #field_sheet.cell(row=rowNum, column=2).value = till_80_val

wb.save('combined_data.xlsx')
print("combined_data.xlsx created\n\n")
