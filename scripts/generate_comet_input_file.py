# import system modules
import os, sys, time

# import functions for extracting data from excel
from openpyxl import load_workbook

# check if argument for excel workbook has been given
if len(sys.argv) < 1:
    print("\n")
    print("python3 generate_comet_input_file.py <spreadsheet location>")
    print("\n")
    # print("Command-line arguments are as follows:")
    # print("  <spreadsheet locatiion> system location of spreadsheet with input data")
    # print("    eg /usr/local/name/comet/data.xml")
    # print("")
    exit()

script_dir = sys.argv[0]
# Get the scenario name, crop id, and iso class to name xml and track
if len(sys.argv) > 1:
    integrated_scenario_name = sys.argv[2]
else:
    integrated_scenario_name = 'no_arg_for_scenario_file_name'

wb_dir = sys.argv[1]
wb = load_workbook(filename = wb_dir)
processed_fields_sheet = wb['processed']
processed_sheets = {}
gis_file_name = 'crp_id_iso_id'
for row in range(1, processed_fields_sheet.max_row + 1):
    sheet_name = processed_fields_sheet['B' + str(row)].value

    scenario_sheet = wb[sheet_name]

    current_values = {}
    # Current practices - years 2000 - 2019
    current_yearly = []
    # Scenario a practices - future years 2020 - 2029
    scenario_yearly = []
    # Scenario b practices - future years 2020 - 2029
    scenario_b_yearly = []

    # loop through scenario values
    for row in range(1, scenario_sheet.max_row + 1):
        if row < 28 or row > 98:
            param  = scenario_sheet['A' + str(row)].value
            param_val  = scenario_sheet['B' + str(row)].value
            current_values.setdefault(param, param_val)
        elif row > 31 and row < 72:
            current_year = {}
            for col in range(2, scenario_sheet.max_column + 5): # no sure why max_column is 18 when it should be more
                year_key = scenario_sheet.cell(row=31, column=col).value
                year_value = scenario_sheet.cell(row=row, column=col).value
                current_year.setdefault(year_key, year_value)
            current_yearly.append(current_year)
        elif row > 75 and row < 96:
            scenario_year = {}
            for col in range(2, scenario_sheet.max_column + 5): # no sure why max_column is 18 when it shjould be more
                year_key = scenario_sheet.cell(row=75, column=col).value
                year_value = scenario_sheet.cell(row=row, column=col).value
                scenario_year.setdefault(year_key, year_value)
            scenario_yearly.append(scenario_year)
        # elif row > 52 and row < 63:
            # scenario_b_year = {}
            # for col in range(1, scenario_sheet.max_column):
                # year_key = scenario_sheet.cell(row=52, column=col).value
                # year_value = scenario_sheet.cell(row=row, column=col).value
                # scenario_b_year.setdefault(year_key, year_value)
            # scenario_b_yearly.append(scenario_b_year)

    current_values.setdefault('yearly_current_data', current_yearly)
    current_values.setdefault('yearly_scenario_data', scenario_yearly)

    scenario_a_name  = str(scenario_sheet['B74'].value)
    # scenario_b_name  = scenario_sheet['B51'].value
    scenario_b_name = '' # no scenario b possible at this time DP 6.10.20
    # if scenario_b_name:
        # current_values.setdefault('yearly_scenario_b_data', scenario_b_yearly)
    # else:
    current_values.setdefault('yearly_scenario_b_data', '')

    # set scenario names
    current_values.setdefault('scenario_a_name', scenario_a_name)
    current_values.setdefault('scenario_b_name', scenario_b_name)

    field_number = sheet_name[6:] # assumes sheet_name is 'ready_field#', removes 'ready_'
    processed_sheets.setdefault(field_number, current_values)

    gis_file_name = scenario_sheet['B105'].value

# create XML file
# timestamp = time.time()

# relative path for mac and linux
script_rel_path = os.path.dirname(os.path.realpath(__file__))
# absolute path for mike on neoterra
script_path = "G:\\projects\\Moore_Climate2018\\COMET-Farm-master\\scripts\\"
master_path = "G:\\projects\\Moore_Climate2018\\COMET-Farm-master\\"

if sys.platform.startswith('darwin') or sys.platform.startswith('linux'):
    input_xml_file = script_rel_path + '/../inputs/api_inputs/' + integrated_scenario_name + '.xml'
elif sys.platform.startswith('win32') or sys.platform.startswith('cygwin'):
    input_xml_file = master_path + 'inputs\\api_inputs\\' + integrated_scenario_name + '.xml'

if (os.path.isfile(input_xml_file)):
    print('Saving over previous version\n')
    print(str(input_xml_file))
    os.remove(input_xml_file)
else:
    print('Creating new XML file')
    print(str(input_xml_file))


# initialize the file content string that will contain the text to be written to the file
with open(input_xml_file, 'w') as f:

    print('Writing XML. This could take a minute...\n')

    # email_address = list(processed_sheets.items())[0][1]['Email']
    f.write("<CometFarm>")
    f.write("<Project ID=\"34505\" PNAME=\"" + str(integrated_scenario_name) + "\" USERID=\"15436\">")
    # f.write("<Day cometEmailId=\"dpollard@ecotrust.org\">")
    for field in processed_sheets:
        #f.write("<Cropland name=\"" + str(integrated_scenario_name) + "\">")
        f.write("<Cropland>")
        #f.write("<GEOM PARCELNAME=\"" + str(processed_sheets[field]['Scenario Name']) + "\" SRID=\"" + str(processed_sheets[field]['SRID']) + "\" AREA=\"" + str(processed_sheets[field]['AREA']) + "\">" + processed_sheets[field]['GEOM'] + "</GEOM>")
        f.write("<GEOM PARCELNAME=\"" + str(processed_sheets[field]['id']) + "\" SRID=\"" + str(processed_sheets[field]['SRID']) + "\" AREA=\"" + str(processed_sheets[field]['AREA']) + "\">" + processed_sheets[field]['GEOM'] + "</GEOM>")
        f.write("<Pre-1980>" + processed_sheets[field]['pre_80'] + "</Pre-1980>")
        f.write("<CRPStartYear>0</CRPStartYear><CRPEndYear>0</CRPEndYear><CRPType>None</CRPType>") # CRP always None
        f.write("<Year1980-2000>" + processed_sheets[field]['yr80_2000'] + "</Year1980-2000>")
        f.write("<Year1980-2000_Tillage>" + processed_sheets[field]['till80_200'] + "</Year1980-2000_Tillage>")

        # start crop scenario
        f.write("<CropScenario Name=\"Current\">")
        for crop_year in processed_sheets[field]['yearly_current_data']: # crop years
            if str(crop_year['Year']) != 'None':
                if str(crop_year['Ccop_name']) != 'None':
                    f.write("<CropYear Year=\"" + str(crop_year['Year']) + "\">")
                    f.write("<Crop CropNumber=\"" + str(crop_year['crop_number']) + "\">")
                    f.write("<CropName>" + str(crop_year['Ccop_name']) + "</CropName>")
                    f.write("<CropType>" + str(crop_year['CropType']) + "</CropType>")
                    f.write("<PlantingDate>" + str(crop_year['planting_date']).strip('\"') + " 07:00:00</PlantingDate>")
                    f.write("<ContinueFromPreviousYear>" + str(crop_year['continue_from_previous_year']) + "</ContinueFromPreviousYear>")

                    # start harvest list
                    f.write("<HarvestList>")
                    f.write("<HarvestEvent>")
                    f.write("<HarvestDate>" + str(crop_year['harvest_date']).strip('\"') + "</HarvestDate>")
                    f.write("<Grain>" + str(crop_year['grain']) + "</Grain>")
                    f.write("<yield>" + str(crop_year['yield']) + "</yield>")
                    f.write("<StrawStoverHayRemoval>" + str(crop_year['straw_stover_hay_removal']) + "</StrawStoverHayRemoval>")
                    f.write("</HarvestEvent>")
                    f.write("</HarvestList>")

                    f.write("<GrazingList />")

                    f.write("<TillageList>")
                    f.write("<TillageEvent>")
                    f.write("<TillageType>" + str(crop_year['tillage_type']) + "</TillageType>")
                    f.write("<TillageDate>" + str(crop_year['tillage_date']).strip('\"') + "</TillageDate>")
                    f.write("</TillageEvent>")
                    f.write("</TillageList>")

                   # start fertilizer list
                    if str(crop_year['n_application_type']) != 'None': #add OMAD tags if compost has been specified
                        f.write("<NApplicationList>") # todo should i add conditional
                        f.write("<NApplicationEvent>")
                        f.write("<NApplicationType>" + str(crop_year['n_application_type']) + "</NApplicationType>")
                        f.write("<NApplicationMethod>" + str(crop_year['n_application_method']) + "</NApplicationMethod>")
                        f.write("<NApplicationDate>" + str(crop_year['n_application_date']).strip('\"') + "</NApplicationDate>")
                        f.write("<NApplicationAmount>" + str(crop_year['n_application_amount']) + "</NApplicationAmount>")
                        f.write("<PApplicationAmount>" + str(crop_year['p_application_amount']) + "</PApplicationAmount>")
                        f.write("<EEP>" + str(crop_year['eep']) + "</EEP>")
                        f.write("</NApplicationEvent>")
                        f.write("</NApplicationList>")
                    else:
                        f.write("<NApplicationList />")


                    if str(crop_year['compost_type']) != 'None': #add OMAD tags if compost has been specified
                        f.write("<OMADApplicationList>")
                        f.write("<OMADApplicationEvent>")
                        f.write("<OMADType>" + str(crop_year['compost_type']) + "</OMADType>")
                        # f.write("<OMADApplicationDate>\"" + str(crop_year['']) + "\"</OMADApplicationDate>")
                        f.write("<OMADApplicationDate>" + str(crop_year['n_application_date']).strip('\"') + "</OMADApplicationDate>")

                        f.write("<OMADAmount>" + str(crop_year['compost_amount']) + "</OMADAmount>")
                        f.write("<OMADPercentN>" + str(crop_year['compost_percent_n']) + "</OMADPercentN>")
                        f.write("<OMADCNRatio>" + str(crop_year['compost_cn_ratio']) + "</OMADCNRatio>")
                        f.write("</OMADApplicationEvent>")
                        f.write("</OMADApplicationList>")
                    else:
                        f.write("<OMADApplicationList />")


                    f.write("<IrrigationList />")

                    f.write("<BurnEvent>")
                    f.write("<BurnTime>No burning</BurnTime>")
                    f.write("</BurnEvent>")

                    f.write("</Crop>")
                    f.write("</CropYear>")

        f.write("</CropScenario>")

        f.write("<CropScenario Name=\"" + processed_sheets[field]['scenario_a_name'] + "\">")

        for crop_year in processed_sheets[field]['yearly_scenario_data']:
             if str(crop_year['Year']) != 'None' or crop_year['Year'] != None:
                c = str(crop_year['Ccop_name']) #check to see if there's a cover crop
                if c != 'None':
                    print (str(crop_year['Ccop_name']) + str(crop_year['Year']))
                    f.write("<CropYear Year=\"" + str(crop_year['Year']) + "\">")
                    f.write("<Crop CropNumber=\"" + str(crop_year['crop_number']) + "\">")
                    f.write("<CropName>" + str(crop_year['Ccop_name']) + "</CropName>")
                    f.write("<CropType>" + str(crop_year['CropType']) + "</CropType>")
                    f.write("<PlantingDate>" + str(crop_year['planting_date']).strip('\"') + " 07:00:00</PlantingDate>")
                    f.write("<ContinueFromPreviousYear>" + str(crop_year['continue_from_previous_year']) + "</ContinueFromPreviousYear>")

    				# start harvest list
                    f.write("<HarvestList>")
                    f.write("<HarvestEvent>")
                    f.write("<HarvestDate>" + str(crop_year['harvest_date']).strip('\"') + "</HarvestDate>")
                    f.write("<Grain>" + str(crop_year['grain']) + "</Grain>")
                    f.write("<yield>" + str(crop_year['yield']) + "</yield>")
                    f.write("<StrawStoverHayRemoval>" + str(crop_year['straw_stover_hay_removal']) + "</StrawStoverHayRemoval>")
                    f.write("</HarvestEvent>")
                    f.write("</HarvestList>")

                    f.write("<GrazingList />")

                    f.write("<TillageList>")
                    f.write("<TillageEvent>")
                    f.write("<TillageType>" + str(crop_year['tillage_type']) + "</TillageType>")
                    f.write("<TillageDate>" + str(crop_year['tillage_date']).strip('\"') + "</TillageDate>")
                    f.write("</TillageEvent>")
                    f.write("</TillageList>")

                    # start fertilizer list
                    if str(crop_year['n_application_type']) != 'None': #add OMAD tags if compost has been specified
                        f.write("<NApplicationList>") # todo should i add conditional
                        f.write("<NApplicationEvent>")
                        f.write("<NApplicationType>" + str(crop_year['n_application_type']) + "</NApplicationType>")
                        f.write("<NApplicationMethod>" + str(crop_year['n_application_method']) + "</NApplicationMethod>")
                        f.write("<NApplicationDate>" + str(crop_year['n_application_date']).strip('\"') + "</NApplicationDate>")
                        f.write("<NApplicationAmount>" + str(crop_year['n_application_amount']) + "</NApplicationAmount>")
                        f.write("<PApplicationAmount>" + str(crop_year['p_application_amount']) + "</PApplicationAmount>")
                        f.write("<EEP>" + str(crop_year['eep']) + "</EEP>")
                        f.write("</NApplicationEvent>")
                        f.write("</NApplicationList>")
                    else:
                        f.write("<NApplicationList />")


                    if str(crop_year['compost_type']) != 'None': #add OMAD tags if compost has been specified
                        f.write("<OMADApplicationList>")
                        f.write("<OMADApplicationEvent>")
                        f.write("<OMADType>" + str(crop_year['compost_type']) + "</OMADType>")
                        # f.write("<OMADApplicationDate>\"" + str(crop_year['']) + "\"</OMADApplicationDate>")
                        f.write("<OMADApplicationDate>" + str(crop_year['n_application_date']).strip('\"') + "</OMADApplicationDate>")

                        f.write("<OMADAmount>" + str(crop_year['compost_amount']) + "</OMADAmount>")
                        f.write("<OMADPercentN>" + str(crop_year['compost_percent_n']) + "</OMADPercentN>")
                        f.write("<OMADCNRatio>" + str(crop_year['compost_cn_ratio']) + "</OMADCNRatio>")
                        f.write("</OMADApplicationEvent>")
                        f.write("</OMADApplicationList>")
                    else:
                        f.write("<OMADApplicationList />")

                    f.write("<IrrigationList />")

                    f.write("<BurnEvent>")
                    f.write("<BurnTime>No burning</BurnTime>")
                    f.write("</BurnEvent>")

                    f.write("</Crop>")
                    f.write("</CropYear>")

        f.write("</CropScenario>")

        # if len(processed_sheets[field]['yearly_scenario_b_data']) > 1:
        #
        #     f.write("<CropScenario Name=\"" + processed_sheets[field]['scenario_b_name'] + "\">")
        #
        #     for crop_year in processed_sheets[field]['yearly_scenario_b_data']:
        #         f.write("<CropYear Year=\"" + str(crop_year['Year']) + "\">")
        #         f.write("<Crop CropNumber=\"1\">")
        #         f.write("<CropName>" + str(crop_year['Ccop_name']) + "</CropName>")
        #         f.write("<CropType>CROPS</CropType>")
        #         f.write("<PlantingDate>" + str(crop_year['planting_date']).strip('\"') + "</PlantingDate>")
        #         f.write("<ContinueFromPreviousYear>" + str(crop_year['continue_from_previous_year']) + "</ContinueFromPreviousYear>")
        #         # f.write("<DidYouPrune></DidYouPrune>") # todo
        #         # f.write("<RenewOrClearYourOrchard></RenewOrClearYourOrchard>") # todo
        #         # start harvest list
        #         f.write("<HarvestList>") # todo should i add conditional
        #         f.write("<HarvestEvent>") # todo
        #         f.write("<HarvestDate>" + str(crop_year['harvest_date']).strip('\"') + "</HarvestDate>")
        #         f.write("<Grain>" + str(crop_year['grain']) + "</Grain>")
        #         f.write("<yield>" + str(crop_year['yield']) + "</yield>")
        #         f.write("<StrawStoverHayRemoval>" + str(crop_year['straw_stover_hay_removal']) + "</StrawStoverHayRemoval>")
        #         f.write("</HarvestEvent>")
        #         f.write("</HarvestList>")
        #         # end harvest list
        #         f.write("<GrazingList />")
        #         f.write("<TillageList>")
        #         f.write("<TillageEvent>")
        #         f.write("<TillageType>" + str(crop_year['tillage_type']) + "</TillageType>")
        #         f.write("<TillageDate>" + str(crop_year['tillage_date']).strip('\"') + "</TillageDate>")
        #         f.write("</TillageEvent>")
        #         f.write("</TillageList>")
        #         # start fertilizer list
        #         f.write("<NApplicationList>") # todo should i add conditional
        #         f.write("<NApplicationEvent>")
        #         f.write("<NApplicationType>" + str(crop_year['n_application_type']) + "</NApplicationType>")
        #         f.write("<NApplicationMethod>" + str(crop_year['n_application_method']) + "</NApplicationMethod>")
        #         f.write("<NApplicationDate>" + str(crop_year['n_application_date']).strip('\"') + "</NApplicationDate>")
        #         f.write("<NApplicationAmount>" + str(crop_year['n_application_amount']) + "</NApplicationAmount>")
        #         f.write("<PApplicationAmount>" + "0" + "</PApplicationAmount>")
        #         f.write("<EEP>" + str(crop_year['eep']) + "</EEP>")
        #         f.write("</NApplicationEvent>")
        #         f.write("</NApplicationList>")
        #         # end fertilizer list
        #         # start omad application list
        #         f.write("<OMADApplicationList />")
        #         # end omad list
        #         # start irrigation list
        #         f.write("<IrrigationList />")
        #         # end irrigation list
        #         # start liming list
        #         f.write("<LimingList>")
        #         f.write("<LimingApplicationEvent>")
        #         f.write("<LimingApplicationDate></LimingApplicationDate>")
        #         f.write("<LimingMaterial></LimingMaterial>")
        #         f.write("<LimingApplicationAmount></LimingApplicationAmount>")
        #         f.write("</LimingApplicationEvent>")
        #         f.write("</LimingList>")
        #         # end liming list
        #         # start burning list
        #         # f.write("<BurningEvent>")
        #         # f.write("<DidYouBurnCropResidue>" + 'No' + "</DidYouBurnCropResidue>")
        #         # f.write("</BurningEvent>")
        #         f.write("<BurnEvent>")
        #         f.write("<BurnTime>No burning</BurnTime>")
        #         f.write("</BurnEvent>")
        #         # end burning list
        #         f.write("</Crop>")
        #         f.write("</CropYear>")
        #
        #     f.write("</CropSscenario>")

        f.write("</Cropland>\n")

    # f.write("</Day>")
    f.write("</Project>")
    f.write("</CometFarm>")

f.close()

print('XML file generated\n')

print('Sendng XML POST request to API')

import requests

if (os.path.isfile(input_xml_file)):
    xml_file = open(input_xml_file, 'rb')
    url = "https://comet-farm.com///ApiMain/AddToQueue"
    data = {
        "LastCropland": -1,
        "FirstCropland": -1,
        "email": "mike@ecotrust.org",
        "LastDaycentInput":0,
        "FirstDaycentInput":0,
        "URL":"",
        "Files":xml_file,
    }
    try:
        req = requests.post(url, data = data)
        print('POST sent to: ' + url)
        print('Response: ' + req.text)
    finally:
        xml_file.close()

else:
    print('\n')
    print('File could not be found')
    print(input_xml_file)
