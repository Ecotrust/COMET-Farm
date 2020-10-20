# import system modules
import os, sys, csv, datetime
from datetime import datetime

# import functions for extracting data from excel
from openpyxl import load_workbook

# First arg should be gis data
#   usually in form of comet_dat.txt
# Second argument for excel workbook has been given
#   will usually be template_crop_id
#   if none given will use template_v2.xlsx within COMET-Farm_Master dir
if len(sys.argv) == 1 or len(sys.argv) > 4:
    print('\nMissing arguments')
    print('expecting 1 or 2 arguments')
    print('  1. GIS data')
    print('  2. spreadsheet v2 (optional)')
    print('\nexpected command (Windows sub `python3` with `py -3`)')
    print('  `python3 ./script/create_api_input.py <GIS data location> <spreadsheet location>`\n')
    print('Command-line arguments are as follows:\n')
    print('  * <GIS data location> system location of comma separated data from GIS')
    print('      e.g.,  /usr/local/name/comet/data.csv\n')
    print('  * <spreadsheet locatiion> system location of spreadsheet to add GIS data')
    print('      e.g.,  /usr/local/name/comet/data.xlsx')
    print('      ** if not given defaults to template_v2.xlsx at repo root\n')
    exit()

script_dir =  sys.argv[0]
gis_dat_file =  sys.argv[1]

gis_file_name = os.path.basename(gis_dat_file)
gis_file_name = gis_file_name.split('.')[0]
gis_file_name = gis_file_name.split('_')
iso_id = gis_file_name[-1]
crop_id = gis_file_name[-2]

if len(sys.argv) > 1:
    wb_dir = sys.argv[2]
    file_name = sys.argv[3]
else:
    script_path = os.path.dirname(os.path.realpath(__file__))
    wb_dir = script_path + '/../template_v3.xlsx'
    file_name = 'no_file_name_provided_as_arg'
    print('No scenario template provided. Using default of template_v3.xlsx')


wb = load_workbook(filename = wb_dir, data_only = True)
scenario_sheet = wb['scenario']

gis_values = {}

# create sheet for list of processed fields
processed_sheet = wb.create_sheet('processed')


# open GIS data file and save it as dict
with open(gis_dat_file) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    csv_dict = csv.DictReader(csv_file)
    gis_values = csv_dict
    run_name = 'crop_id_and_iso_id_not_in_gis_file_name'

    if len(crop_id) > 0 and len(iso_id) > 0:
        run_name = crop_id + '_' + iso_id

    # for row in gis_values:
        # print(dict(row))
        # param  = scenario_sheet['A' + str(row)].value
        # param_val  = scenario_sheet['B' + str(row)].value
        # scenario_values.setdefault(param, param_val)

    # add gis data to spreadsheet
        # field_ID,CRP,CRPType,GEOM,SRID,AREA,pre_80,yr80_2000,CRP_NUM,CcopName,planting_date,harvest_date,grain,till_date,n_app_date
    for row in gis_values:
        field_sheet = wb.copy_worksheet(scenario_sheet)
        field_sheet.title = 'ready_' + row['field_ID']
        processed_sheet.append(['name', field_sheet.title])

        if 'pre_80' in row:
            value_to_match = int(row['pre_80'])
            match_value = wb['pre1980'].cell(row=value_to_match, column=1).value
            field_sheet.cell(row=100, column=2).value = str(match_value)
        if 'yr80_2000' in row:
            value_to_match = int(row['yr80_2000'])
            match_value = wb['yr80'].cell(row=value_to_match, column=1).value
            field_sheet.cell(row=101, column=2).value = str(match_value)
        if 'till80_200' in row:
            value_to_match = int(row['till80_200'])
            match_value = wb['tillage'].cell(row=value_to_match, column=1).value
            field_sheet.cell(row=102, column=2).value = str(match_value)
        if 'crop_scenario_name' in row:
            field_sheet.cell(row=105, column=2).value = row['crop_scenario_name']
        else:
            field_sheet.cell(row=105, column=2).value = run_name
        if 'CRP_NUM' in row:
            field_sheet.cell(row=106, column=2).value = row['CRP_NUM'] #dc added 1/16/20
        if 'field_ID' in row:
            field_sheet.cell(row=107, column=2).value = row['field_ID']
        if 'GEOM' in row:
            field_sheet.cell(row=108, column=2).value = 'POLYGON (' + row['GEOM'] + ')'
        if 'AREA' in row:
            field_sheet.cell(row=109, column=2).value = row['AREA']
        if 'SRID' in row:
            field_sheet.cell(row=110, column=2).value = row['SRID']
        if 'Scenario Name' in row:
            field_sheet.cell(row=1, column=2).value = row['Scenario Name']
        else:
            field_sheet.cell(row=1, column=2).value = run_name

        for crop_cell in field_sheet.iter_cols(min_col=3,max_col=23,min_row=32,max_row=71):
            for cell in crop_cell:
                if cell.row % 2 == 0:
                    if str(cell.column) == 'C' or cell.column == 3:
                        if 'Ccop_name' in row:
                            cell.value = row['Ccop_name']
                    elif str(cell.column) == 'D' or str(cell.column) == 'F' or str(cell.column) == 'K' or str(cell.column) == 'M' or cell.column == 4 or cell.column == 6 or cell.column == 11 or cell.column == 13:
                        # get year from template spreadsheet
                        yyyy = field_sheet.cell(row=cell.row, column=2).value
                        # get month and day from GIS data
                        # formated as number day of year ex: 32
                        if str(cell.column) == 'D' or cell.column == 4:
                            if 'planting_date' in row:
                                day_num_of_year = row['planting_date']
                        elif str(cell.column) == 'F' or cell.column == 6:
                            if 'harvest_date' in row:
                                day_num_of_year = row['harvest_date']
                        elif str(cell.column) == 'K' or cell.column == 11:
                            if 'till_date' in row:
                                day_num_of_year = row['till_date']
                        elif str(cell.column) == 'M' or cell.column == 13:
                            if 'n_app_date' in row:
                                day_num_of_year = row['n_app_date']

                        #-**- See Day of year chart at end of document

                        # combine and format
                        month_day = datetime.strptime(str(day_num_of_year), '%j')
                        # convert to the orginially expected format of Month, day ex: March 14
                        month_day = month_day.strftime('%B %d')
                        # add the year to month day from the template spreadsheet
                        mmddyyyy = month_day + str(yyyy)
                        # format ex: March 152020
                        mmddyyyy = datetime.strptime(mmddyyyy, '%B %d%Y')
                        # convert to CF API expected format ex: 03/14/2020
                        cfarm_format_date = mmddyyyy.strftime('%m/%d/%Y')
                        # add formated date to template spreadsheet
                        cell.value = cfarm_format_date

                elif cell.row % 2 == 1:
                    cover_crop_name = field_sheet.cell(row=4, column=2).value
                    if str(cover_crop_name) != 'None':
                        if str(cell.column) == 'C' or cell.column == 3:
                            if 'Ccop_name' in row:
                                cell.value = row['Ccop_name']
                        # elif str(cell.column) == 'D' or str(cell.column) == 'F' or str(cell.column) == 'K' or str(cell.column) == 'M':
                        elif str(cell.column) == 'D' or str(cell.column) == 'F' or cell.column == 4 or cell.column == 6:

                            yyyy = field_sheet.cell(row=cell.row, column=2).value

                            # formated as number day of year ex: 32
                            if str(cell.column) == 'D' or cell.column == 4:
                                if 'planting_date' in row:
                                    date_to_parse = row['planting_date']
                                    diff_column_needed = 6 # The column we will need later for setting cover crop date
                                    date_diff = 5 # if we are planting a cover crop we want to do so 5 days after the previous crop harvest date
                                    diff_year = 0 # planting a cover crop - we want to know previous crop
                            elif str(cell.column) == 'F' or cell.column == 6:
                                if 'harvest_date' in row:
                                    date_to_parse = row['harvest_date']
                                    diff_column_needed = 4 # The column we will need later for setting cover crop date
                                    date_diff = -5 # if we are harvesting a cover crop we want to do so 5 days before planting next crop
                                    diff_year = 1 # about to plant a new crop - we want to know the next crop
                            # elif cell.column == 'K':
                            #     if 'till_date' in row:
                            #         date_to_parse = row['till_date']
                            # elif cell.column == 'M':
                            #     if 'n_app_date' in row:
                            #         date_to_parse = row['n_app_date']

                            diff_row_num = cell.row + diff_year # neg or pos 1 depending on harverst or planting
                            diff_crop_date = field_sheet.cell(row=diff_row_num, column=diff_column_needed) # get the crop date
                            yyyy = yyyy + diff_year
                            diff_crop_date_val = diff_crop_date.value
                            if diff_crop_date_val:
                                diff_crop_date_val = diff_crop_date_val.replace(''','')
                                diff_crop_date_datetime = datetime.strptime(str(diff_crop_date_val), '%m/%d/%Y') # formate to day number year so we can do some math
                                diff_crop_date_to_day_num = diff_crop_date_datetime.strftime('%j')

                            five_days_diff = int(diff_crop_date_to_day_num) + date_diff # pos or neg day of year diff
                            five_days_diff = datetime.strptime(str(five_days_diff), '%j') # convert back to datetime
                            #-**- See Day of year chart at end of document

                            # convert back to CF needed format
                            # convert to the orginially expected format of Month, day ex: March 14
                            month_day = five_days_diff.strftime('%B %d')
                            # get the year from crop (not cover crop)
                            # yyyy = diff_crop_date[-5:-1]
                            # add the year to month day from the template spreadsheet
                            mmddyyyy = month_day + str(yyyy)
                            # format ex: March 152020
                            mmddyyyy = datetime.strptime(mmddyyyy, '%B %d%Y')
                            # convert to CF API expected format ex: 03/14/2020
                            cfarm_format_date = mmddyyyy.strftime('%m/%d/%Y')
                            # add formated date to template spreadsheet
                            cell.value = cfarm_format_date

        for crop_cell in field_sheet.iter_cols(min_col=3,max_col=23,min_row=76,max_row=95):
            for cell in crop_cell:
                if cell.row % 2 == 0:
                    if str(cell.column) == 'C' or cell.column == 3:
                        if 'Ccop_name' in row:
                            cell.value = row['Ccop_name']
                    elif str(cell.column) == 'D' or str(cell.column) == 'F' or str(cell.column) == 'K' or str(cell.column) == 'M' or cell.column == 4 or cell.column == 6 or cell.column == 11 or cell.column == 13:
                        # get year from template spreadsheet
                        yyyy = field_sheet.cell(row=cell.row, column=2).value
                        # get month and day from GIS data
                        # formated as number day of year ex: 32
                        if str(cell.column) == 'D' or cell.column == 4:
                            if 'planting_date' in row:
                                day_num_of_year = row['planting_date']
                        elif str(cell.column) == 'F' or cell.column == 6:
                            if 'harvest_date' in row:
                                day_num_of_year = row['harvest_date']
                        elif str(cell.column) == 'K' or cell.column == 11:
                            if 'till_date' in row:
                                day_num_of_year = row['till_date']
                        elif str(cell.column) == 'M' or cell.column == 13:
                            if 'n_app_date' in row:
                                day_num_of_year = row['n_app_date']

                        #-**- See Day of year chart at end of document

                        # combine and format
                        month_day = datetime.strptime(str(day_num_of_year), '%j')
                        # convert to the orginially expected format of Month, day ex: March 14
                        month_day = month_day.strftime('%B %d')
                        # add the year to month day from the template spreadsheet
                        mmddyyyy = month_day + str(yyyy)
                        # format ex: March 152020
                        mmddyyyy = datetime.strptime(mmddyyyy, '%B %d%Y')
                        # convert to CF API expected format ex: 03/14/2020
                        cfarm_format_date = mmddyyyy.strftime('%m/%d/%Y')
                        # add formated date to template spreadsheet
                        cell.value = cfarm_format_date

                elif cell.row % 2 == 1:
                    cover_crop_name = field_sheet.cell(row=18, column=2).value
                    if str(cover_crop_name) != 'None':
                        if str(cell.column) == 'C' or cell.column == 3:
                            if 'Ccop_name' in row:
                                cell.value = row['Ccop_name']
                        # elif str(cell.column) == 'D' or str(cell.column) == 'F' or str(cell.column) == 'K' or str(cell.column) == 'M':
                        elif str(cell.column) == 'D' or str(cell.column) == 'F' or cell.column == 4 or cell.column == 6:

                            yyyy = field_sheet.cell(row=cell.row, column=2).value

                            # formated as number day of year ex: 32
                            if str(cell.column) == 'D' or cell.column == 4:
                                if 'planting_date' in row:
                                    date_to_parse = row['planting_date']
                                    diff_column_needed = 6 # The column we will need later for setting cover crop date
                                    date_diff = 5 # if we are planting a cover crop we want to do so 5 days after the previous crop harvest date
                                    diff_year = 0 # planting a cover crop - we want to know previous crop
                            elif str(cell.column) == 'F' or cell.column == 6:
                                if 'harvest_date' in row:
                                    date_to_parse = row['harvest_date']
                                    diff_column_needed = 4 # The column we will need later for setting cover crop date
                                    date_diff = -5 # if we are harvesting a cover crop we want to do so 5 days before planting next crop
                                    diff_year = 1 # about to plant a new crop - we want to know the next crop
                            # elif cell.column == 'K':
                            #     if 'till_date' in row:
                            #         date_to_parse = row['till_date']
                            # elif cell.column == 'M':
                            #     if 'n_app_date' in row:
                            #         date_to_parse = row['n_app_date']

                            diff_row_num = cell.row + diff_year # neg or pos 1 depending on harverst or planting
                            diff_crop_date = field_sheet.cell(row=diff_row_num, column=diff_column_needed) # get the crop date
                            yyyy = yyyy + diff_year
                            diff_crop_date_val = diff_crop_date.value
                            if diff_crop_date_val:
                                diff_crop_date_val = diff_crop_date_val.replace(''','')
                                diff_crop_date_datetime = datetime.strptime(str(diff_crop_date_val), '%m/%d/%Y') # formate to day number year so we can do some math
                                diff_crop_date_to_day_num = diff_crop_date_datetime.strftime('%j')

                            five_days_diff = int(diff_crop_date_to_day_num) + date_diff # pos or neg day of year diff
                            five_days_diff = datetime.strptime(str(five_days_diff), '%j') # convert back to datetime
                            #-**- See Day of year chart at end of document

                            # convert back to CF needed format
                            # convert to the orginially expected format of Month, day ex: March 14
                            month_day = five_days_diff.strftime('%B %d')
                            # get the year from crop (not cover crop)
                            # yyyy = diff_crop_date[-5:-1]
                            # add the year to month day from the template spreadsheet
                            mmddyyyy = month_day + str(yyyy)
                            # format ex: March 152020
                            mmddyyyy = datetime.strptime(mmddyyyy, '%B %d%Y')
                            # convert to CF API expected format ex: 03/14/2020
                            cfarm_format_date = mmddyyyy.strftime('%m/%d/%Y')
                            # add formated date to template spreadsheet
                            cell.value = cfarm_format_date

print('Successful integration\n')

# relative path for mac and linux
script_rel_path = os.path.dirname(os.path.realpath(__file__))
# absolute path for mike on neoterra
script_path = 'G:\\projects\\Moore_Climate2018\\COMET-Farm-master\\scripts\\'
master_path = 'G:\\projects\\Moore_Climate2018\\COMET-Farm-master\\'

if sys.platform.startswith('darwin') or sys.platform.startswith('linux'):
    wb.save(master_path + '/integrated/integrated_' + file_name + '_' + crop_id + '_' + iso_id + '.xlsx')
    # os.system('python3 ' + script_rel_path + '/generate_comet_input_file.py' + ' ' + script_rel_path + '/../combined_data.xlsx')
elif sys.platform.startswith('win32') or sys.platform.startswith('cygwin'):
    # Save using scenario name, crop id, and iso class
    wb.save(master_path + '\\integrated\\integrated_' + file_name + '_' + crop_id + '_' + iso_id + '.xlsx')


################################################################
##******     NOTES                      ******************* ####
################################################################

# Day of Year Chart
# First column within each month is for regular years. The second is for leap years: 1972, 1976, 1980, 1984, 1988, 1992, 1996, 2000, 2004, 2008, 2012, 2016, 2020.
#
# day of
# month	Jan	Feb	Mar	Apr	May	Jun	Jul	Aug	Sep	Oct	Nov	Dec
# 1	001	032 032	060 061	091 092	121 122	152 153	182 183	213 214	244 245	274 275	305 306	335 336
# 2	002	033 033	061 062	092 093	122 123	153 154	183 184	214 215	245 246	275 276	306 307	336 337
# 3	003	034 034	062 063	093 094	123 124	154 155	184 185	215 216	246 247	276 277	307 308	337 338
# 4	004	035 035	063 064	094 095	124 125	155 156	185 186	216 217	247 248	277 278	308 309	338 339
# 5	005	036 036	064 065	095 096	125 126	156 157	186 187	217 218	248 249	278 279	309 310	339 340
# 6	006	037 037	065 066	096 097	126 127	157 158	187 188	218 219	249 250	279 280	310 311	340 341
# 7	007	038 038	066 067	097 098	127 128	158 159	188 189	219 220	250 251	280 281	311 312	341 342
# 8	008	039 039	067 068	098 099	128 129	159 160	189 190	220 221	251 252	281 282	312 313	342 343
# 9	009	040 040	068 069	099 100	129 130	160 161	190 191	221 222	252 253	282 283	313 314	343 344
# 10	010	041 041	069 070	100 101	130 131	161 162	191 192	222 223	253 254	283 284	314 315	344 345
# 11	011	042 042	070 071	101 102	131 132	162 163	192 193	223 224	254 255	284 285	315 316	345 346
# 12	012	043 043	071 072	102 103	132 133	163 164	193 194	224 225	255 256	285 286	316 317	346 347
# 13	013	044 044	072 073	103 104	133 134	164 165	194 195	225 226	256 257	286 287	317 318	347 348
# 14	014	045 045	073 074	104 105	134 135	165 166	195 196	226 227	257 258	287 288	318 319	348 349
# 15	015	046 046	074 075	105 106	135 136	166 167	196 197	227 228	258 259	288 289	319 320	349 350
# 16	016	047 047	075 076	106 107	136 137	167 168	197 198	228 229	259 260	289 290	320 321	350 351
# 17	017	048 048	076 077	107 108	137 138	168 169	198 199	229 230	260 261	290 291	321 322	351 352
# 18	018	049 049	077 078	108 109	138 139	169 170	199 200	230 231	261 262	291 292	322 323	352 353
# 19	019	050 050	078 079	109 110	139 140	170 171	200 201	231 232	262 263	292 293	323 324	353 354
# 20	020	051 051	079 080	110 111	140 141	171 172	201 202	232 233	263 264	293 294	324 325	354 355
# 21	021	052 052	080 081	111 112	141 142	172 173	202 203	233 234	264 265	294 295	325 326	355 356
# 22	022	053 053	081 082	112 113	142 143	173 174	203 204	234 235	265 266	295 296	326 327	356 357
# 23	023	054 054	082 083	113 114	143 144	174 175	204 205	235 236	266 267	296 297	327 328	357 358
# 24	024	055 055	083 084	114 115	144 145	175 176	205 206	236 237	267 268	297 298	328 329	358 359
# 25	025	056 056	084 085	115 116	145 146	176 177	206 207	237 238	268 269	298 299	329 330	359 360
# 26	026	057 057	085 086	116 117	146 147	177 178	207 208	238 239	269 270	299 300	330 331	360 361
# 27	027	058 058	086 087	117 118	147 148	178 179	208 209	239 240	270 271	300 301	331 332	361 362
# 28	028	059 059	087 088	118 119	148 149	179 180	209 210	240 241	271 272	301 302	332 333	362 363
# 29	029	      060	088 089	119 120	149 150	180 181	210 211	241 242	272 273	302 303	333 334	363 364
# 30	030	 	089 090	120 121	150 151	181 182	211 212	242 243	273 274	303 304	334 335	364 365
# 31	031	 	090 091	 	151 152	 	212 213	243 244	 	304 305	 	365 366
